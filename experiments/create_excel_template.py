#!/usr/bin/env python3
"""
Script to create Excel template for Fibonacci trading journal
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_fibonacci_journal_template():
    """Create Excel template for Fibonacci trading journal"""

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fibonacci Trading Journal"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column headers
    headers = [
        ("A", "№", 5),
        ("B", "Дата/Время", 16),
        ("C", "Пара", 12),
        ("D", "Таймфрейм", 12),
        ("E", "Тип свечи", 12),
        ("F", "High свечи", 12),
        ("G", "Low свечи", 12),
        ("H", "Диапазон", 12),
        ("I", "Entry #1 (0.0)", 14),
        ("J", "Entry #2 (0.5)", 14),
        ("K", "Entry #3 (0.618)", 14),
        ("L", "Stop Loss (0.82)", 14),
        ("M", "TP1 (-0.618)", 14),
        ("N", "TP2 (-1.618)", 14),
        ("O", "TP3 (-2.618)", 14),
        ("P", "Размер позиции (%)", 16),
        ("Q", "Статус", 12),
        ("R", "Прибыль/Убыток", 14),
        ("S", "Комментарии", 30),
    ]

    # Set column widths and headers
    for col_letter, header_text, width in headers:
        ws.column_dimensions[col_letter].width = width
        cell = ws[f"{col_letter}1"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = "A2"

    # Add example rows
    examples = [
        {
            "date": "2026-02-04 10:00",
            "pair": "ETHUSDT",
            "tf": "1H",
            "type": "🟢 LONG",
            "high": 93500,
            "low": 93000,
            "range": 500,
            "entry1": 93500,
            "entry2": 93250,
            "entry3": 93191,
            "stop": 93090,
            "tp1": 93809,
            "tp2": 94309,
            "tp3": 94809,
            "size": "3%",
            "status": "Открыта",
            "pnl": "",
            "comment": "Пример из Issue #97",
        },
        {
            "date": "2026-02-04 14:00",
            "pair": "BTCUSDT",
            "tf": "4H",
            "type": "🔴 SHORT",
            "high": 95000,
            "low": 94500,
            "range": 500,
            "entry1": 94500,
            "entry2": 94750,
            "entry3": 94809,
            "stop": 94910,
            "tp1": 94191,
            "tp2": 93691,
            "tp3": 93191,
            "size": "3%",
            "status": "Закрыта",
            "pnl": "+450",
            "comment": "Пример SHORT позиции",
        },
    ]

    for idx, example in enumerate(examples, start=2):
        row = idx
        ws[f"A{row}"] = idx - 1
        ws[f"B{row}"] = example["date"]
        ws[f"C{row}"] = example["pair"]
        ws[f"D{row}"] = example["tf"]
        ws[f"E{row}"] = example["type"]
        ws[f"F{row}"] = example["high"]
        ws[f"G{row}"] = example["low"]
        ws[f"H{row}"] = example["range"]
        ws[f"I{row}"] = example["entry1"]
        ws[f"J{row}"] = example["entry2"]
        ws[f"K{row}"] = example["entry3"]
        ws[f"L{row}"] = example["stop"]
        ws[f"M{row}"] = example["tp1"]
        ws[f"N{row}"] = example["tp2"]
        ws[f"O{row}"] = example["tp3"]
        ws[f"P{row}"] = example["size"]
        ws[f"Q{row}"] = example["status"]
        ws[f"R{row}"] = example["pnl"]
        ws[f"S{row}"] = example["comment"]

        # Apply styles
        for col in range(1, 20):  # A to S
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment if col <= 2 else left_alignment

        # Color coding
        if example["type"] == "🟢 LONG":
            ws[f"E{row}"].fill = green_fill
        else:
            ws[f"E{row}"].fill = red_fill

        ws[f"I{row}"].fill = green_fill if example["type"] == "🟢 LONG" else red_fill
        ws[f"J{row}"].fill = blue_fill
        ws[f"K{row}"].fill = blue_fill
        ws[f"L{row}"].fill = red_fill
        ws[f"M{row}"].fill = yellow_fill
        ws[f"N{row}"].fill = yellow_fill
        ws[f"O{row}"].fill = yellow_fill

    # Add empty rows for user input (rows 4-20)
    for row in range(4, 21):
        ws[f"A{row}"] = row - 1
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 1:
                cell.alignment = center_alignment

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Инструкция")
    ws_instructions.column_dimensions['A'].width = 80

    instructions = [
        "📊 ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ЖУРНАЛА FIBONACCI TRADING",
        "",
        "1. ПОДГОТОВКА",
        "   - Откройте TradingView и перейдите на график нужной пары (например, ETHUSDT)",
        "   - Выберите таймфрейм (например, 1 час)",
        "   - Добавьте индикатор 'Fibonacci Trading Calculator' на график",
        "",
        "2. ВЫБОР СВЕЧИ",
        "   - В индикаторе установите параметр 'Lookback Bars' = 0 для текущей свечи",
        "   - Или установите = 1, 2, 3... для предыдущих свечей",
        "   - Индикатор автоматически определит тип свечи (🟢 зеленая или 🔴 красная)",
        "",
        "3. СБОР ДАННЫХ",
        "   - Посмотрите на таблицу индикатора справа вверху графика",
        "   - Скопируйте следующие данные в Excel:",
        "     * Дата/Время - текущая дата и время",
        "     * Пара - название торговой пары (ETHUSDT, BTCUSDT и т.д.)",
        "     * Таймфрейм - выбранный таймфрейм (1H, 4H, 1D и т.д.)",
        "     * Тип свечи - 🟢 LONG или 🔴 SHORT",
        "     * High свечи - максимальная цена",
        "     * Low свечи - минимальная цена",
        "     * Диапазон - разница между High и Low",
        "     * Entry #1 (0.0) - первый вход по рынку",
        "     * Entry #2 (0.5) - второй вход (лимитный ордер)",
        "     * Entry #3 (0.618) - третий вход (лимитный ордер)",
        "     * Stop Loss (0.82) - стоп-лосс для всех входов",
        "     * TP1 (-0.618) - первый тейк-профит (закрыть 30%)",
        "     * TP2 (-1.618) - второй тейк-профит (закрыть 30%)",
        "     * TP3 (-2.618) - третий тейк-профит (закрыть 40%)",
        "",
        "4. УПРАВЛЕНИЕ ПОЗИЦИЕЙ",
        "   - Entry #1: Вход по рынку при закрытии сигнальной свечи (1% депозита)",
        "   - Entry #2: Лимитный ордер на уровне 0.5 (1% депозита)",
        "   - Entry #3: Лимитный ордер на уровне 0.618 (1% депозита)",
        "   - Stop Loss: Защитный стоп на уровне 0.82 для всех входов",
        "   - При достижении TP1: закрыть 30% позиции",
        "   - При достижении TP2: закрыть еще 30% позиции",
        "   - При достижении TP3: закрыть оставшиеся 40% позиции",
        "",
        "5. ОТСЛЕЖИВАНИЕ РЕЗУЛЬТАТОВ",
        "   - В колонке 'Статус' отмечайте: Открыта, TP1, TP2, TP3, Закрыта, Stop Loss",
        "   - В колонке 'Прибыль/Убыток' записывайте финальный результат сделки",
        "   - В колонке 'Комментарии' добавляйте важные заметки",
        "",
        "6. ЦВЕТОВАЯ КОДИРОВКА",
        "   - 🟢 Зеленый - LONG позиции и Entry #1",
        "   - 🔴 Красный - SHORT позиции и Stop Loss",
        "   - 🔵 Синий - Entry #2 и Entry #3",
        "   - 🟡 Желтый - Take Profit уровни",
        "",
        "7. ПРИМЕР ИСПОЛЬЗОВАНИЯ",
        "   См. первые две строки в журнале (примеры LONG и SHORT)",
        "",
        "8. ПОЛЕЗНЫЕ ССЫЛКИ",
        "   - Репозиторий: https://github.com/alekseymavai/TRADERAGENT",
        "   - Issue #97: https://github.com/alekseymavai/TRADERAGENT/issues/97",
        "",
        "📝 ВАЖНО:",
        "   - Всегда используйте Stop Loss для защиты капитала",
        "   - Не рискуйте более 3% депозита на одну сделку (1% × 3 входа)",
        "   - Ведите дневник регулярно для анализа своих результатов",
        "   - Тестируйте стратегию на демо-счете перед реальной торговлей",
        "",
        "Успешной торговли! 🚀",
    ]

    for idx, instruction in enumerate(instructions, start=1):
        cell = ws_instructions[f"A{idx}"]
        cell.value = instruction
        cell.alignment = left_alignment

        if instruction.startswith("📊"):
            cell.font = Font(bold=True, size=14, color="1F4E78")
        elif instruction and instruction[0].isdigit():
            cell.font = Font(bold=True, size=12)
        elif instruction.startswith("📝"):
            cell.font = Font(bold=True, size=11, color="C00000")

    # Save workbook
    output_file = "FIBONACCI_TRADING_JOURNAL_TEMPLATE.xlsx"
    wb.save(output_file)
    print(f"✓ Excel template created: {output_file}")
    return output_file


if __name__ == "__main__":
    try:
        output_file = create_fibonacci_journal_template()
        print(f"\n✓ Success! Template saved as: {output_file}")
        print("\nTemplate includes:")
        print("  - Main journal sheet with formatted columns")
        print("  - Example LONG and SHORT trades")
        print("  - Color-coded cells for easy reading")
        print("  - Instructions sheet in Russian")
    except Exception as e:
        print(f"✗ Error creating template: {e}")
        import traceback
        traceback.print_exc()
