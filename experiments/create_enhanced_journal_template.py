#!/usr/bin/env python3
"""
Создание расширенного шаблона журнала сделок с информацией о свечах
Enhanced trading journal template with candle information (OHLC, levels, etc.)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_enhanced_journal_template():
    """Создает расширенный Excel шаблон для журнала сделок"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trading Journal"

    # Определяем стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    # Стиль для положительных значений
    profit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    profit_font = Font(color="006100", bold=True)

    # Стиль для отрицательных значений
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    loss_font = Font(color="9C0006", bold=True)

    # Стиль для границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки колонок
    headers = [
        "Trade #",           # A
        "Дата и Время",      # B
        "Пара",              # C
        "Таймфрейм",         # D
        "Направление",       # E
        # Информация о свече
        "Свеча Open",        # F
        "Свеча High",        # G
        "Свеча Low",         # H
        "Свеча Close",       # I
        "Размер свечи (%)",  # J
        # Fibonacci уровни
        "Entry #1",          # K
        "Entry #2",          # L
        "Entry #3",          # M
        "Stop Loss",         # N
        "TP1",               # O
        "TP2",               # P
        "TP3",               # Q
        # Исполнение
        "Entry #1 Filled",   # R
        "Entry #2 Filled",   # S
        "Entry #3 Filled",   # T
        "TP1 Hit",           # U
        "TP2 Hit",           # V
        "TP3 Hit",           # W
        "Stop Hit",          # X
        # Результаты
        "Profit/Loss ($)",   # Y
        "Profit/Loss (%)",   # Z
        "Комментарий"        # AA
    ]

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Примеры данных (LONG сделка - успешная)
    example_long_success = [
        1,                           # Trade #
        "2024-02-04 10:00",         # Дата
        "ETHUSDT",                  # Пара
        "1h",                       # Таймфрейм
        "LONG",                     # Направление
        93000,                      # Open
        93500,                      # High
        93000,                      # Low
        93500,                      # Close
        "=((G2-H2)/H2)*100",       # Размер свечи % (формула)
        93500,                      # Entry #1
        93250,                      # Entry #2
        93191,                      # Entry #3
        93090,                      # Stop Loss
        93809,                      # TP1
        94309,                      # TP2
        94809,                      # TP3
        "✓",                        # Entry #1 Filled
        "✓",                        # Entry #2 Filled
        "✓",                        # Entry #3 Filled
        "✓",                        # TP1 Hit
        "✓",                        # TP2 Hit
        "✓",                        # TP3 Hit
        "✗",                        # Stop Hit
        "+450.00",                  # Profit $
        "+4.83%",                   # Profit %
        "Все TP достигнуты"         # Комментарий
    ]

    # Примеры данных (SHORT сделка - стоп)
    example_short_stop = [
        2,                           # Trade #
        "2024-02-05 14:00",         # Дата
        "BTCUSDT",                  # Пара
        "1h",                       # Таймфрейм
        "SHORT",                    # Направление
        42000,                      # Open
        42500,                      # High
        42000,                      # Low
        42200,                      # Close
        "=((G3-H3)/H3)*100",       # Размер свечи %
        42000,                      # Entry #1
        42250,                      # Entry #2
        42309,                      # Entry #3
        42410,                      # Stop Loss
        41691,                      # TP1
        40691,                      # TP2
        39691,                      # TP3
        "✓",                        # Entry #1 Filled
        "✓",                        # Entry #2 Filled
        "✗",                        # Entry #3 Filled
        "✗",                        # TP1 Hit
        "✗",                        # TP2 Hit
        "✗",                        # TP3 Hit
        "✓",                        # Stop Hit
        "-120.00",                  # Profit $
        "-1.27%",                   # Profit %
        "Сработал Stop Loss"        # Комментарий
    ]

    # Примеры данных (LONG сделка - частичный успех)
    example_long_partial = [
        3,                           # Trade #
        "2024-02-10 09:30",         # Дата
        "ETHUSDT",                  # Пара
        "4h",                       # Таймфрейм
        "LONG",                     # Направление
        2800,                       # Open
        2850,                       # High
        2800,                       # Low
        2830,                       # Close
        "=((G4-H4)/H4)*100",       # Размер свечи %
        2850,                       # Entry #1
        2825,                       # Entry #2
        2819,                       # Entry #3
        2809,                       # Stop Loss
        2881,                       # TP1
        2931,                       # TP2
        2981,                       # TP3
        "✓",                        # Entry #1 Filled
        "✓",                        # Entry #2 Filled
        "✓",                        # Entry #3 Filled
        "✓",                        # TP1 Hit
        "✗",                        # TP2 Hit
        "✗",                        # TP3 Hit
        "✗",                        # Stop Hit
        "+85.00",                   # Profit $
        "+0.91%",                   # Profit %
        "TP1 достигнут, остальное в безубыток"
    ]

    # Записываем примеры
    for row_num, row_data in enumerate([example_long_success, example_short_stop, example_long_partial], 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border

            # Выравнивание
            if col_num in [1, 5, 18, 19, 20, 21, 22, 23, 24]:  # Trade #, Direction, checkmarks
                cell.alignment = center_align

            # Форматирование прибыли/убытка
            if col_num == 25:  # Profit/Loss $
                if isinstance(value, str) and value.startswith('+'):
                    cell.fill = profit_fill
                    cell.font = profit_font
                elif isinstance(value, str) and value.startswith('-'):
                    cell.fill = loss_fill
                    cell.font = loss_font

    # Настройка ширины колонок
    column_widths = {
        'A': 10,   # Trade #
        'B': 18,   # Дата
        'C': 12,   # Пара
        'D': 12,   # Таймфрейм
        'E': 14,   # Направление
        'F': 12,   # Open
        'G': 12,   # High
        'H': 12,   # Low
        'I': 12,   # Close
        'J': 16,   # Размер %
        'K': 12,   # Entry #1
        'L': 12,   # Entry #2
        'M': 12,   # Entry #3
        'N': 12,   # Stop Loss
        'O': 12,   # TP1
        'P': 12,   # TP2
        'Q': 12,   # TP3
        'R': 14,   # Entry #1 Filled
        'S': 14,   # Entry #2 Filled
        'T': 14,   # Entry #3 Filled
        'U': 12,   # TP1 Hit
        'V': 12,   # TP2 Hit
        'W': 12,   # TP3 Hit
        'X': 12,   # Stop Hit
        'Y': 14,   # Profit $
        'Z': 14,   # Profit %
        'AA': 30,  # Комментарий
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Замораживаем первую строку
    ws.freeze_panes = "A2"

    # Добавляем инструкции на второй лист
    ws_instructions = wb.create_sheet("Инструкция")

    instructions = [
        ["РАСШИРЕННЫЙ ШАБЛОН ЖУРНАЛА СДЕЛОК FIBONACCI"],
        [""],
        ["Этот шаблон содержит дополнительные колонки для информации о свечах:"],
        [""],
        ["ИНФОРМАЦИЯ О СВЕЧЕ:"],
        ["- Свеча Open/High/Low/Close: OHLC данные выбранной свечи"],
        ["- Размер свечи (%): Диапазон свечи в процентах от цены"],
        [""],
        ["FIBONACCI УРОВНИ:"],
        ["- Entry #1, #2, #3: Рассчитанные уровни входа"],
        ["- Stop Loss: Уровень остановки убытков (0.820)"],
        ["- TP1, TP2, TP3: Уровни взятия прибыли"],
        [""],
        ["ИСПОЛНЕНИЕ:"],
        ["- Entry #1/2/3 Filled: Был ли вход исполнен (✓ или ✗)"],
        ["- TP1/2/3 Hit: Достигнут ли уровень Take Profit"],
        ["- Stop Hit: Сработал ли Stop Loss"],
        [""],
        ["КАК ЗАПОЛНЯТЬ:"],
        [""],
        ["ВАРИАНТ 1: Вручную (из индикатора)"],
        ["1. Откройте Fibonacci Calculator в TradingView"],
        ["2. Посмотрите информационную таблицу с уровнями"],
        ["3. Скопируйте OHLC свечи и уровни в этот шаблон"],
        ["4. Отметьте какие ордера исполнены"],
        ["5. Запишите результат (прибыль/убыток)"],
        [""],
        ["ВАРИАНТ 2: Из стратегии (автоматически)"],
        ["1. Запустите Fibonacci Auto Strategy в TradingView"],
        ["2. Откройте Strategy Tester -> List of Trades"],
        ["3. Экспортируйте в CSV"],
        ["4. Импортируйте данные в Excel"],
        ["5. Дополните OHLC данные из маркеров на графике"],
        [""],
        ["СТАТИСТИКА:"],
        [""],
        ["Вы можете добавить сводную таблицу для анализа:"],
        ["- Всего сделок"],
        ["- Успешных сделок (TP3)"],
        ["- Стопов (Stop Loss)"],
        ["- Средняя прибыль"],
        ["- Win Rate (%)"],
        ["- Profit Factor"],
        [""],
        ["СОВЕТЫ:"],
        [""],
        ["1. Используйте фильтры Excel для анализа по направлению (LONG/SHORT)"],
        ["2. Сортируйте по прибыли чтобы найти лучшие/худшие сделки"],
        ["3. Группируйте по торговой паре для сравнения"],
        ["4. Анализируйте корреляцию между размером свечи и результатом"],
        ["5. Сохраняйте резервные копии журнала!"],
        [""],
        ["Дата создания: 04.02.2024"],
        ["Версия: 2.0 (Enhanced with Candle Info)"],
    ]

    for row_num, row_data in enumerate(instructions, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num)
            cell.value = value

            # Форматирование заголовков
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="4472C4")
            elif value and isinstance(value, str) and value.endswith(":") and len(value) < 50:
                cell.font = Font(bold=True, size=11)

    ws_instructions.column_dimensions['A'].width = 80

    # Сохраняем файл
    filename = "FIBONACCI_ENHANCED_JOURNAL_TEMPLATE.xlsx"
    wb.save(filename)
    print(f"✅ Расширенный шаблон создан: {filename}")
    print(f"📊 Колонки: {len(headers)}")
    print(f"📝 Примеров сделок: 3")
    print(f"📖 Страниц: {len(wb.sheetnames)}")

    return filename

if __name__ == "__main__":
    create_enhanced_journal_template()
